from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict

def create_calendar(weeks: Dict[str, List[Dict[str, str]]], filename: str = "calendar.xlsx"):
    # Create a new workbook
    wb = Workbook()
    
    for week_num, days in weeks.items():
        # Create a new worksheet for each week
        if week_num == '1':
            ws = wb.active
            ws.title = f"Week {week_num}"
        else:
            ws = wb.create_sheet(f"Week {week_num}")

        # Set the width of the columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Set the height of the rows
        for row in range(1, 7):
            ws.row_dimensions[row].height = 50

        # Create a list of days of the week
        days_of_week = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

        # Add the days of the week to the first row of the worksheet
        for col, day in enumerate(days_of_week, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = day
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="00FFFF")

        # Apply a light blue background color and borders to the cells below the days of the week
        for row in range(2, 7):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(fill_type="solid", fgColor="ADD8E6")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Add events to the calendar based on the extracted information from the text
        for day in days:
            day_num = int(day['day'])
            start_time = day['start_time']
            event_name = day['event_name']
            ws.cell(row=day_num+1, column=day_num).value = f"{event_name} at {start_time}"
        
        # Add "(Rest)" to empty events
        for row in range(2, 7):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    cell.value = "(Rest)"

    # Save the workbook to a file
    wb.save(filename)
